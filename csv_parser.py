import csv
from openpyxl import Workbook
import operator


class Scrapper():

    def __init__(self):
        self.salary_filename = input('Введите идеальный путь к файлу vacancies_dif_currency.csv: ')

    def parse_salary(self):
        with open(self.salary_filename, newline='', encoding='utf-8') as file:
            profdata = {}
            reader = csv.reader(file, delimiter=',', quotechar='|')
            worklst = ['qa', 'test', 'тест', 'quality assurance']
            try:
                for row in reader:
                    try:
                        for profession in worklst:
                            if profession in row[0].lower():
                                year = row[-1].split('-')[0]
                                if row[1] != '' or row[2] != '':
                                    if row[1] != '':
                                        salary = row[1]
                                    else:
                                        salary = row[2]
                                    try:
                                        profdata[f"{year}_count"] = profdata[f"{year}_count"] + 1
                                        profdata[year] = profdata[year] + int(salary.split('.')[0])
                                    except Exception:
                                        profdata[year] = int(salary.split('.')[0])
                                        profdata[f"{year}_count"] = 1
                                    continue
                    except Exception:
                        pass
            except Exception:
                pass
        with open(self.salary_filename, newline='', encoding='utf-8') as file:
            data = {}
            reader = csv.reader(file, delimiter=',', quotechar='|')
            try:
                for row in reader:
                    try:
                        year = row[-1].split('-')[0]
                        if row[0] == 'name':
                            continue
                        if row[1] != '' or row[2] != '':
                            if row[1] != '':
                                salary = row[1]
                            else:
                                salary = row[2]
                            try:
                                data[f"{year}_count"] = data[f"{year}_count"] + 1
                                data[year] = data[year] + int(salary.split('.')[0])
                            except Exception:
                                data[year] = int(salary.split('.')[0])
                                data[f"{year}_count"] = 1
                            continue
                    except Exception:
                        continue
            except Exception:
                pass
        prof_salary = {}
        salary = {}
        prof_work = {}
        work = {}
        for key, value in profdata.items():
            if 'count' in key:
                prof_work[key.split('_')[0]] = profdata[key]
                continue
            prof_salary[key] = profdata[key] // profdata[f"{key}_count"]
        for key, value in data.items():
            if 'count' in key:
                work[key.split('_')[0]] = data[key]
                continue
            salary[key] = data[key] // data[f"{key}_count"]

        return {
            'prof_salary': prof_salary,
            'salary': salary,
            'prof_work': prof_work,
            'work': work
        }

    def parse_cities(self):
        with open(self.salary_filename, newline='', encoding='utf-8') as file:
            reader = csv.reader(file, delimiter=',', quotechar='|')
            worklst = ['qa', 'test', 'тест', 'quality assurance']
            data = {}
            try:
                for row in reader:
                    try:
                        for profession in worklst:
                            if profession in row[0].lower():
                                city = row[-2]
                                if row[1] != '' or row[2] != '':
                                    if row[3] != 'USD':
                                        continue
                                    if row[1] != '':
                                        salary = row[1]
                                    else:
                                        salary = row[2]
                                    try:
                                        data[city] = data[city] + int(salary.split('.')[0])
                                        data[f"{city}_count"] = data[f"{city}_count"] + 1
                                    except Exception:
                                        data[city] = int(salary.split('.')[0])
                                        data[f"{city}_count"] = 1
                                continue
                    except Exception:
                        continue
            except Exception:
                pass
        salary = {}
        vacancy = {}
        for key, value in data.items():
            if "count" in key:
                vacancy[key.split('_')[0]] = value
                continue
            salary[key] = data[key] // data[f"{key}_count"]
        return {
            'vacancy': vacancy,
            'cities': salary
        }

    def create_new_file(self, data: dict):
        wb = Workbook()
        filename = 'qa_tester.xlsx'
        ws1 = wb.active
        ws1.title = "QA tester"
        i = 3

        parse_dict = {
            'prof_salary': 'Уровень зарплат по годам тестировщика',
            'salary': 'Уровень зарплат по годам',
            'prof_work': 'Количество вакансий по годам тестировщика',
            'work': 'Количество вакансий по годам',
            'vacancy': 'Доля вакансий по городам',
            'cities': 'Уровень зарплат по городам'

        }

        for main_key, main_values in data.items():
            for title, values in main_values.items():
                for key_title, value_title in parse_dict.items():
                    if title == key_title:
                        break
                ws1[f"B{i}"] = value_title
                i += 2
                col = 3
                for point_title, point_value in values.items():
                    cell = ws1.cell(row=i, column=col)
                    cell.value = point_title
                    cell = ws1.cell(row=(i + 1), column=col)
                    cell.value = point_value
                    col += 1
                i += 3

        wb.save(filename=filename)


    def generate_info(self):
        salary = self.parse_salary()
        cities = self.parse_cities()

        # sort
        for key, value in cities.items():
            sorted_dict = dict(sorted(value.items(), key=operator.itemgetter(1), reverse=True))
            cities[key] = sorted_dict

        self.create_new_file({'salary': salary, "cities": cities})


if __name__ == '__main__':
    scrap = Scrapper()
    scrap.generate_info()
